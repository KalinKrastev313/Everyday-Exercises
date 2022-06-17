# If you have a column of string values of the type "{grams} g"
# the following code transforms these values into "{ounces} oz" in the next column.

from openpyxl import load_workbook, Workbook

file_to_transform = "test.xlsx"
row_input = "A"
row_output = "B"
wb = load_workbook(file_to_transform)
ws = wb.active


for row in range(1, 10):
    if "g" in ws[row_input + str(row)].value:
        grams = ""
        for char in ws[row_input + str(row)].value:
            if char.isdigit() or ".":
                grams += char
            else:
                break
        ws[row_output + str(row)].value = f"{(float(grams) * 0.035274):.2f} oz"
    else:
        ws[row_output + str(row)].value = ws[row_input + str(row)].value


wb.save(file_to_transform)